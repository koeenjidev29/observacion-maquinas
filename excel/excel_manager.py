import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

class ExcelManager:
    def __init__(self, excel_path="data/observaciones.xlsx"):
        self.excel_path = excel_path
        self.workbook = None
        self.load_or_create_workbook()
    
    def load_or_create_workbook(self):
        """Carga el archivo Excel existente o crea uno nuevo"""
        if os.path.exists(self.excel_path):
            try:
                self.workbook = openpyxl.load_workbook(self.excel_path)
                # Verificar que hay al menos una hoja
                if len(self.workbook.sheetnames) == 0:
                    # Si no hay hojas, crear una hoja inicial
                    today = datetime.now().strftime("%Y-%m-%d")
                    self.create_date_sheet(today)
            except Exception as e:
                # Si hay error al cargar, crear nuevo workbook
                self.workbook = Workbook()
                # Renombrar la hoja por defecto con la fecha de hoy
                today = datetime.now().strftime("%d-%m-%Y")
                self.workbook.active.title = today
                self._setup_sheet_headers(self.workbook.active)
                self.save_workbook()
        else:
            self.workbook = Workbook()
            # Renombrar la hoja por defecto con la fecha de hoy
            today = datetime.now().strftime("%d-%m-%Y")
            self.workbook.active.title = today
            self._setup_sheet_headers(self.workbook.active)
            self.save_workbook()
    
    def get_sheet_name_from_date(self, date):
        """Convierte una fecha en nombre de pestaña"""
        if isinstance(date, str):
            # Lista de formatos de fecha soportados
            date_formats = [
                "%d/%m/%Y",    # dd/mm/yyyy
                "%Y-%m-%d",    # yyyy-mm-dd  
                "%m/%d/%y",    # m/d/yy (formato americano corto)
                "%m/%d/%Y",    # m/d/yyyy (formato americano)
                "%d/%m/%y"     # dd/mm/yy
            ]
            
            for date_format in date_formats:
                try:
                    parsed_date = datetime.strptime(date, date_format)
                    # Si el año es de 2 dígitos y menor a 50, asumimos 20xx
                    if parsed_date.year < 1950:
                        parsed_date = parsed_date.replace(year=parsed_date.year + 2000)
                    return parsed_date.strftime("%d-%m-%Y")
                except ValueError:
                    continue
            
            # Si ningún formato funciona, mostrar error descriptivo
            raise ValueError(f"Formato de fecha no válido: {date}. Formatos soportados: dd/mm/yyyy, yyyy-mm-dd, m/d/yy, m/d/yyyy")
        
        return date.strftime("%d-%m-%Y")
    
    def _setup_sheet_headers(self, worksheet):
        """Configura los encabezados de una hoja"""
        headers = ["ID", "Hora", "Línea", "Máquina", "Observación", "Usuario"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar ancho de columnas
        worksheet.column_dimensions['A'].width = 8   # ID
        worksheet.column_dimensions['B'].width = 12  # Hora
        worksheet.column_dimensions['C'].width = 15  # Línea
        worksheet.column_dimensions['D'].width = 20  # Máquina
        worksheet.column_dimensions['E'].width = 50  # Observación
        worksheet.column_dimensions['F'].width = 15  # Usuario
    
    def sheet_exists(self, sheet_name):
        """Verifica si existe una pestaña con el nombre dado"""
        return sheet_name in self.workbook.sheetnames
    
    def create_date_sheet(self, date):
        """Crea una nueva pestaña para la fecha especificada"""
        sheet_name = self.get_sheet_name_from_date(date)
        
        if self.sheet_exists(sheet_name):
            return self.workbook[sheet_name]
        
        # Crear nueva pestaña
        worksheet = self.workbook.create_sheet(title=sheet_name)
        
        # Configurar encabezados usando el método auxiliar
        self._setup_sheet_headers(worksheet)
        
        self.save_workbook()
        return worksheet
    
    def add_observation(self, date, machine, observation, user="Usuario"):
        """Añade una observación a la pestaña de la fecha especificada"""
        sheet_name = self.get_sheet_name_from_date(date)
        
        # Crear pestaña si no existe
        if not self.sheet_exists(sheet_name):
            worksheet = self.create_date_sheet(date)
        else:
            worksheet = self.workbook[sheet_name]
        
        # Encontrar la siguiente fila vacía
        next_row = worksheet.max_row + 1
        
        # Añadir datos
        current_time = datetime.now().strftime("%H:%M:%S")
        worksheet.cell(row=next_row, column=1, value=current_time)
        worksheet.cell(row=next_row, column=2, value=machine)
        worksheet.cell(row=next_row, column=3, value=observation)
        worksheet.cell(row=next_row, column=4, value=user)
        
        # Aplicar formato a las celdas
        for col in range(1, 5):
            cell = worksheet.cell(row=next_row, column=col)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        self.save_workbook()
        return True
    
    def get_observations_by_date(self, date):
        """Obtiene todas las observaciones de una fecha específica"""
        sheet_name = self.get_sheet_name_from_date(date)
        
        if not self.sheet_exists(sheet_name):
            return []
        
        worksheet = self.workbook[sheet_name]
        observations = []
        
        # Leer datos desde la fila 2 (saltando encabezados)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # Si hay datos en la fila
                # Verificar si tenemos el formato nuevo (con línea) o el antiguo
                if len(row) >= 5:  # Formato nuevo: Hora, Línea, Máquina, Observación, Usuario
                    hora_formateada = self._format_time(row[0])
                    observations.append((
                        date,           # 0: fecha
                        hora_formateada, # 1: hora
                        row[1] or "",   # 2: línea
                        row[2] or "",   # 3: máquina
                        row[3] or "",   # 4: observación
                        row[4] or "Usuario"  # 5: usuario
                    ))
                else:  # Formato antiguo: Hora, Máquina, Observación, Usuario
                    machine_name = row[1] if row[1] else ""
                    line_name = self._extract_line_from_machine(machine_name)
                    hora_formateada = self._format_time(row[0])
                    
                    observations.append((
                        date,           # 0: fecha
                        hora_formateada, # 1: hora
                        line_name,      # 2: línea (extraída)
                        machine_name,   # 3: máquina
                        row[2] or "",   # 4: observación
                        row[3] or "Usuario"  # 5: usuario
                    ))
        
        return observations
    
    def _format_time(self, time_value):
        """Formatea correctamente el valor de tiempo"""
        if time_value is None:
            return "--:--:--"
        
        # Si es un objeto datetime.time
        if hasattr(time_value, 'strftime'):
            return time_value.strftime("%H:%M:%S")
        
        # Si es un string, intentar parsearlo
        if isinstance(time_value, str):
            # Si ya está en formato correcto, devolverlo
            if len(time_value.split(':')) >= 2:
                return time_value
            # Si es solo números, intentar formatear
            try:
                if len(time_value) == 4:  # HHMM
                    return f"{time_value[:2]}:{time_value[2:]}:00"
                elif len(time_value) == 6:  # HHMMSS
                    return f"{time_value[:2]}:{time_value[2:4]}:{time_value[4:]}"
            except:
                pass
        
        # Si es un número (tiempo de Excel)
        if isinstance(time_value, (int, float)):
            try:
                # Convertir tiempo de Excel a horas, minutos, segundos
                total_seconds = int(time_value * 24 * 3600)
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            except:
                pass
        
        # Si no se puede formatear, convertir a string
        return str(time_value)
    
    def _extract_line_from_machine(self, machine_name):
        """Extrae el nombre de la línea basándose en el nombre de la máquina"""
        from utils.helpers import get_lines_list, get_machines_by_line
        
        lines = get_lines_list()
        for line in lines:
            machines = get_machines_by_line(line)
            if machine_name in machines:
                return line
        
        return "Sin línea"  # Valor por defecto si no se encuentra
    
    def get_all_dates(self):
        """Obtiene todas las fechas (pestañas) disponibles"""
        dates = []
        for sheet_name in self.workbook.sheetnames:
            try:
                # Convertir nombre de pestaña a fecha
                date = datetime.strptime(sheet_name, "%d-%m-%Y")
                dates.append(date.strftime("%Y-%m-%d"))
            except ValueError:
                continue  # Ignorar pestañas que no sean fechas
        return sorted(dates)
    
    def get_available_dates(self):
        """Alias para get_all_dates - obtiene todas las fechas disponibles"""
        return self.get_all_dates()
    
    def _normalize_date_format(self, date_str):
        """Convierte fecha de formato DD/MM/YYYY a YYYY-MM-DD"""
        if isinstance(date_str, str):
            # Si ya está en formato YYYY-MM-DD, devolverlo
            if len(date_str) == 10 and date_str[4] == '-' and date_str[7] == '-':
                return date_str
            # Si está en formato DD/MM/YYYY, convertirlo
            elif len(date_str) == 10 and date_str[2] == '/' and date_str[5] == '/':
                day, month, year = date_str.split('/')
                return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        return date_str
    
    def delete_observation(self, date, time, machine, observation):
        """Elimina una observación específica"""
        # Normalizar formato de fecha
        normalized_date = self._normalize_date_format(date)
        sheet_name = self.get_sheet_name_from_date(normalized_date)
        
        if not self.sheet_exists(sheet_name):
            raise Exception(f"No existe la hoja para la fecha {date}")
        
        worksheet = self.workbook[sheet_name]
        
        # Buscar la fila que coincida con los datos
        for row_num in range(2, worksheet.max_row + 1):
            row_data = [worksheet.cell(row=row_num, column=col).value for col in range(1, 5)]
            
            # Comparar hora, máquina y observación
            if (self._format_time(row_data[0]) == time and 
                row_data[1] == machine and 
                row_data[2] == observation):
                
                # Eliminar la fila
                worksheet.delete_rows(row_num)
                self.save_workbook()
                return
        
        raise Exception("No se encontró la observación para eliminar")
    
    def update_observation(self, date, time, old_machine, old_observation, new_machine, new_observation):
        """Actualiza una observación específica"""
        # Normalizar formato de fecha
        normalized_date = self._normalize_date_format(date)
        sheet_name = self.get_sheet_name_from_date(normalized_date)
        
        if not self.sheet_exists(sheet_name):
            raise Exception(f"No existe la hoja para la fecha {date}")
        
        worksheet = self.workbook[sheet_name]
        
        # Buscar la fila que coincida con los datos originales
        for row_num in range(2, worksheet.max_row + 1):
            row_data = [worksheet.cell(row=row_num, column=col).value for col in range(1, 5)]
            
            # Comparar hora, máquina y observación originales
            if (self._format_time(row_data[0]) == time and 
                row_data[1] == old_machine and 
                row_data[2] == old_observation):
                
                # Actualizar los datos
                worksheet.cell(row=row_num, column=2, value=new_machine)
                worksheet.cell(row=row_num, column=3, value=new_observation)
                self.save_workbook()
                return
        
        raise Exception("No se encontró la observación para actualizar")
    
    def migrate_observations_add_user(self, default_user="Produccion1"):
        """Migra observaciones existentes sin usuario asignándoles un usuario por defecto"""
        try:
            migrated_count = 0
            for sheet_name in self.workbook.sheetnames:
                worksheet = self.workbook[sheet_name]
                
                # Verificar si la hoja tiene la estructura correcta
                if worksheet.max_row < 2:  # Solo encabezados o vacía
                    continue
                    
                # Revisar cada fila de datos
                for row_num in range(2, worksheet.max_row + 1):
                    # Verificar si la columna de usuario (columna 4) está vacía
                    user_cell = worksheet.cell(row=row_num, column=4)
                    if user_cell.value is None or str(user_cell.value).strip() == "":
                        # Asignar usuario por defecto
                        user_cell.value = default_user
                        migrated_count += 1
            
            if migrated_count > 0:
                self.save_workbook()
                
            return migrated_count
            
        except Exception as e:
            raise Exception(f"Error al migrar observaciones: {str(e)}")
    
    def clear_all_observations(self):
        """Borra todas las observaciones del archivo Excel"""
        try:
            # Crear nuevo workbook
            self.workbook = Workbook()
            # Crear hoja con fecha actual
            today = datetime.now().strftime("%d-%m-%Y")
            self.workbook.active.title = today
            self._setup_sheet_headers(self.workbook.active)
            self.save_workbook()
            return True
        except Exception as e:
            print(f"Error al borrar observaciones: {str(e)}")
            return False
    
    def get_all_observations_with_id(self):
        """Obtiene todas las observaciones con ID de todas las hojas"""
        all_observations = []
        try:
            for sheet_name in self.workbook.sheetnames:
                worksheet = self.workbook[sheet_name]
                
                for row in range(2, worksheet.max_row + 1):
                    # Verificar si hay datos en la fila
                    if worksheet.cell(row=row, column=2).value:  # Si hay hora
                        # ✅ Convertir ID a entero, con valor por defecto 0
                        id_value = worksheet.cell(row=row, column=1).value
                        try:
                            observation_id = int(id_value) if id_value is not None else 0
                        except (ValueError, TypeError):
                            observation_id = 0
                        
                        observation = {
                            'id': observation_id,  # ✅ Siempre será int
                            'fecha': sheet_name,
                            'hora': self._format_time(worksheet.cell(row=row, column=2).value),
                            'linea': worksheet.cell(row=row, column=3).value or '',
                            'maquina': worksheet.cell(row=row, column=4).value or '',
                            'observacion': worksheet.cell(row=row, column=5).value or '',
                            'usuario': worksheet.cell(row=row, column=6).value or 'Usuario'
                        }
                        all_observations.append(observation)
                # ✅ Mover el ordenamiento y return FUERA del bucle
                all_observations.sort(key=lambda x: x['id'])
                return all_observations
                    
        except Exception as e:
            print(f"Error al obtener observaciones: {str(e)}")
            return []
        
    def save_observation(self, date, time, line, machine, observation, user="Usuario"):
        """Guarda una observación en el archivo Excel"""
        try:
            # Obtener ID automático
            observation_id = self.get_next_id()
            
            sheet_name = self.get_sheet_name_from_date(date)
            
            # Crear pestaña si no existe
            if not self.sheet_exists(sheet_name):
                worksheet = self.create_date_sheet(date)
            else:
                worksheet = self.workbook[sheet_name]
            
            # Encontrar la siguiente fila vacía
            next_row = worksheet.max_row + 1
            
            # Añadir datos CON ID
            worksheet.cell(row=next_row, column=1, value=observation_id)  # ID
            worksheet.cell(row=next_row, column=2, value=time)           # Hora
            worksheet.cell(row=next_row, column=3, value=line)           # Línea
            worksheet.cell(row=next_row, column=4, value=machine)        # Máquina
            worksheet.cell(row=next_row, column=5, value=observation)    # Observación
            worksheet.cell(row=next_row, column=6, value=user)           # Usuario
            
            # Aplicar formato a las celdas
            for col in range(1, 7):
                cell = worksheet.cell(row=next_row, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            self.save_workbook()
            return True
            
        except Exception as e:
            print(f"Error al guardar observación: {str(e)}")
            return False
    
    def save_workbook(self):
        """Guarda el archivo Excel"""
        try:
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
            self.workbook.save(self.excel_path)
        except Exception as e:
            raise Exception(f"Error al guardar el archivo Excel: {str(e)}")
    
    def close(self):
        """Cierra el archivo Excel"""
        if self.workbook:
            self.workbook.close()
    
    def get_next_id(self):
        """Genera el siguiente ID automático consultando todas las observaciones"""
        try:
            max_id = 0
            # Revisar todas las hojas del workbook
            for sheet_name in self.workbook.sheetnames:
                worksheet = self.workbook[sheet_name]
                # Revisar todas las filas (empezar desde la fila 2, saltando headers)
                for row in range(2, worksheet.max_row + 1):
                    # Si hay datos en la fila (verificar si hay hora), incrementar contador
                    if worksheet.cell(row=row, column=1).value:  # Si hay hora
                        max_id += 1
            return max_id + 1
        except Exception as e:
            print(f"Error al generar ID: {str(e)}")
            return 1
    
    def open_excel_file(self):
        """Abre el archivo Excel con la aplicación predeterminada del sistema"""
        try:
            import subprocess
            import os
            
            if os.path.exists(self.excel_path):
                # Para Windows
                if os.name == 'nt':
                    os.startfile(self.excel_path)
                # Para macOS
                elif os.name == 'posix' and os.uname().sysname == 'Darwin':
                    subprocess.call(['open', self.excel_path])
                # Para Linux
                else:
                    subprocess.call(['xdg-open', self.excel_path])
                return True
            else:
                return False
        except Exception as e:
            print(f"Error al abrir Excel: {str(e)}")
            return False