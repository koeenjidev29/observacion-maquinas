#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para crear la plantilla de máquinas en Excel
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_machine_template():
    """Crea la plantilla de máquinas en Excel"""
    
    # Crear nuevo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Máquinas"
    
    # Datos de las máquinas
    machines_data = [
        ["ID", "Nombre de Máquina", "Área", "Responsable", "Estado", "Observaciones"],
        ["A001", "Máquina A - Producción", "Producción", "Juan Pérez", "Activa", "Máquina principal de producción"],
        ["B002", "Máquina B - Ensamblaje", "Ensamblaje", "María García", "Activa", "Línea de ensamblaje automático"],
        ["C003", "Máquina C - Empaque", "Empaque", "Carlos López", "Activa", "Empacadora automática"],
        ["D004", "Máquina D - Control Calidad", "Calidad", "Ana Martínez", "Activa", "Equipo de control de calidad"],
        ["E005", "Máquina E - Mantenimiento", "Mantenimiento", "Luis Rodríguez", "Activa", "Herramientas de mantenimiento"],
        ["F006", "Máquina F - Logística", "Logística", "Carmen Sánchez", "Activa", "Sistema logístico"],
        ["G007", "Máquina G - Almacén", "Almacén", "Roberto Torres", "Activa", "Gestión de almacén"],
        ["H008", "Máquina H - Transporte", "Transporte", "Elena Vargas", "Activa", "Sistema de transporte interno"]
    ]
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Llenar datos
    for row_idx, row_data in enumerate(machines_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            if row_idx == 1:  # Encabezados
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:  # Datos
                cell.alignment = data_alignment
    
    # Ajustar ancho de columnas
    column_widths = [8, 25, 15, 20, 12, 30]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    # Crear segunda hoja con información adicional
    ws2 = wb.create_sheet("Información")
    
    info_data = [
        ["INFORMACIÓN DE LA PLANTILLA DE MÁQUINAS"],
        [""],
        ["Descripción:"],
        ["Esta plantilla contiene la información de todas las máquinas"],
        ["disponibles en el sistema de observación."],
        [""],
        ["Campos:"],
        ["• ID: Identificador único de la máquina"],
        ["• Nombre: Nombre descriptivo de la máquina"],
        ["• Área: Área de trabajo donde se encuentra"],
        ["• Responsable: Persona encargada de la máquina"],
        ["• Estado: Estado actual (Activa/Inactiva/Mantenimiento)"],
        ["• Observaciones: Notas adicionales sobre la máquina"],
        [""],
        ["Uso:"],
        ["Esta plantilla es utilizada por el programa de observación"],
        ["para mostrar las máquinas disponibles en la interfaz."],
        [""],
        ["Fecha de creación:", "2025"],
        ["Versión:", "1.0"]
    ]
    
    for row_idx, row_data in enumerate(info_data, 1):
        if isinstance(row_data, list) and len(row_data) > 0:
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Título
                    cell.font = Font(bold=True, size=14)
                elif "•" in str(value):  # Elementos de lista
                    cell.font = Font(size=10)
    
    # Ajustar ancho de columnas en la segunda hoja
    ws2.column_dimensions['A'].width = 50
    ws2.column_dimensions['B'].width = 15
    
    # Guardar archivo
    file_path = "plantilla_maquinas.xlsx"
    wb.save(file_path)
    print(f"Plantilla creada: {file_path}")
    
    return file_path

if __name__ == "__main__":
    create_machine_template()